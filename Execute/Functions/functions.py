from flask import request, jsonify, session, render_template, redirect, send_file
from Execute import queries
from datetime import date
import io
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter


# =====================================================
# COMMON RESPONSE HELPERS
# =====================================================

def success_response(message="", data=None, status=200):
    res = {"success": True, "message": message}
    if data is not None:
        res["data"] = data
    return jsonify(res), status


def error_response(message="Something went wrong", status=400):
    return jsonify({"success": False, "message": message}), status


# =====================================================
# DROPDOWN APIs
# =====================================================

def get_locations_fn():
    try:
        locs = queries.get_all_locations()
        return success_response(data=locs)
    except Exception as e:
        return error_response(str(e), 500)


def get_months_fn():
    try:
        location = request.args.get("location")
        data     = queries.get_months_by_location(location)
        return success_response(data=data)
    except Exception as e:
        return error_response(str(e), 500)


def get_organizations_by_filter_fn():
    try:
        location   = request.args.get("location")
        month_year = request.args.get("month_year")
        orgs       = queries.get_orgs(location, month_year)
        return success_response(data=orgs)
    except Exception as e:
        return error_response(str(e), 500)


# =====================================================
# SHARED: BUILD REPORT DATA
# =====================================================

def _build_report(year, month, location, organization):

    success, result = queries.fetch_attendance_data(
        year, month, location, organization
    )

    if not success:
        return None, result 

    if not result:
        return None, "No data found for the selected filters."

    available_days = sorted(set(
        rec["PDate"].day if hasattr(rec["PDate"], "day")
        else int(str(rec["PDate"])[8:10])
        for rec in result
    ))

    day_names = [date(year, month, d).strftime("%A") for d in available_days]

    employees = {}
    for rec in result:
        uid = rec["UserID"]
        if uid not in employees:
            employees[uid] = {
                "name":        rec["FullName"],
                "designation": rec["DsgName"],
                "days":        {},
                "ot_total":    0.0
            }

        day_no = rec["PDate"].day if hasattr(rec["PDate"], "day") \
            else int(str(rec["PDate"])[8:10])

        employees[uid]["days"][day_no] = rec["AttendanceValue"]

        try:
            employees[uid]["ot_total"] += float(rec["Overtime"] or 0)
        except (TypeError, ValueError):
            pass

    rows = []
    sr   = 1
    for uid, emp in employees.items():
        day_values = []
        mandays    = 0

        for d in available_days:
            val = emp["days"].get(d, "AB")
            if val is None or str(val).strip() == "":
                val = "AB"
            day_values.append(str(val).strip())
            if val not in ("AB", "WO", "", None):
                mandays += 1

        rows.append({
            "sr_no":       sr,
            "name":        emp["name"],
            "designation": emp["designation"],
            "days":        day_values,
            "mandays":     mandays,
            "ot":          round(emp["ot_total"], 2)
        })
        sr += 1

    return {
        "location":     location,
        "organization": organization,
        "month_label":  date(year, month, 1).strftime("%b-%y"),
        "days":         available_days,
        "day_names":    day_names,
        "rows":         rows
    }, None


# =====================================================
# GENERATE REPORT  (JSON for browser rendering)
# =====================================================

def generate_report_fn():
    try:
        data         = request.get_json()
        month_year   = data.get("month_year")
        location     = data.get("location")
        organization = data.get("organization")

        year  = int(month_year.split("-")[0])
        month = int(month_year.split("-")[1])

        report, err = _build_report(year, month, location, organization)

        if err:
            return error_response(err)

        return success_response(data=report)

    except Exception as e:
        return error_response(str(e), 500)


# =====================================================
# DOWNLOAD EXCEL  (same visual format as screenshot)
# =====================================================

def download_excel_fn():
    try:
        month_year   = request.args.get("month_year")
        location     = request.args.get("location")
        organization = request.args.get("organization")

        year  = int(month_year.split("-")[0])
        month = int(month_year.split("-")[1])

        report, err = _build_report(year, month, location, organization)
        if err:
            return error_response(err)

        wb = _build_excel(report)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = (
            f"Attendance_{location}_{report['month_label'].replace('-','_')}.xlsx"
        )

        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        return error_response(str(e), 500)


# =====================================================
# EXCEL BUILDER  — matches screenshot layout exactly
# =====================================================

def _build_excel(report):
    """
    Build an openpyxl Workbook that exactly mirrors the Excel screenshot:

    Row 1 (merged):  Location | <loc>   Date | <month>  Contractor Name | <org>
    Row 2: blank
    Row 3: Sr No | Name of Employees | Designation | 1 | 2 | ... | N | Mandays | OT | Remarks
    Row 4+: data rows
    """

    # ---- colour / style constants ----
    BLUE_FILL   = PatternFill("solid", fgColor="1565C0")   # header bg
    WHITE_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=9)
    BOLD_FONT   = Font(name="Arial", bold=True, size=9)
    NORMAL_FONT = Font(name="Arial", size=9)
    RED_FONT    = Font(name="Arial", bold=True, color="FF0000", size=9)
    GREY_FONT   = Font(name="Arial", color="888888", size=8, italic=True)

    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    thin = Side(style="thin", color="CCCCCC")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance Report"

    days      = report["days"]
    day_names = report["day_names"]
    rows      = report["rows"]
    n_days    = len(days)

    COL_SR    = 1
    COL_NAME  = 2
    COL_DESIG = 3
    COL_D1    = 4                    
    COL_LAST_DAY = COL_D1 + n_days - 1
    COL_MANDAYS  = COL_LAST_DAY + 1
    COL_OT       = COL_LAST_DAY + 2
    COL_REMARKS  = COL_LAST_DAY + 3
    TOTAL_COLS   = COL_REMARKS

    ws.cell(1, 2, "Location").font   = BOLD_FONT
    ws.cell(1, 3, report["location"]).font = NORMAL_FONT
    ws.cell(1, 4, "Date").font       = BOLD_FONT
    ws.cell(1, 5, report["month_label"]).font = NORMAL_FONT
    ws.cell(1, 6, "Contractor Name").font = BOLD_FONT

    ws.merge_cells(
        start_row=1, start_column=7,
        end_row=1,   end_column=min(16, TOTAL_COLS)
    )
    ws.cell(1, 7, report["organization"]).font = BOLD_FONT
    ws.cell(1, 7).alignment = LEFT

    HDR_ROW = 3

    def hdr(col, val):
        c = ws.cell(HDR_ROW, col, val)
        c.fill      = BLUE_FILL
        c.font      = WHITE_FONT
        c.alignment = CENTER
        c.border    = BORDER

    hdr(COL_SR,    "Sr. No.")
    hdr(COL_NAME,  "Name of Employees")
    hdr(COL_DESIG, "Designation")

    for i, (d, dn) in enumerate(zip(days, day_names)):
        col = COL_D1 + i
        c   = ws.cell(HDR_ROW, col, d)
        c.fill      = BLUE_FILL
        c.font      = WHITE_FONT
        c.alignment = CENTER
        c.border    = BORDER

    hdr(COL_MANDAYS, "Mandays")
    hdr(COL_OT,      "OT")
    hdr(COL_REMARKS, "Remarks")

    DAY_ROW = 4

    ws.cell(DAY_ROW, COL_SR,    "").fill = BLUE_FILL
    ws.cell(DAY_ROW, COL_NAME,  "").fill = BLUE_FILL
    ws.cell(DAY_ROW, COL_DESIG, "").fill = BLUE_FILL
    ws.cell(DAY_ROW, COL_MANDAYS, "").fill = BLUE_FILL
    ws.cell(DAY_ROW, COL_OT,    "").fill = BLUE_FILL
    ws.cell(DAY_ROW, COL_REMARKS,"").fill = BLUE_FILL

    for i, dn in enumerate(day_names):
        col = COL_D1 + i
        c   = ws.cell(DAY_ROW, col, dn[:3])
        c.fill      = PatternFill("solid", fgColor="1E88E5")
        c.font      = Font(name="Arial", bold=False, color="FFFFFF", size=8)
        c.alignment = CENTER
        c.border    = BORDER

    for row_idx, row in enumerate(rows):
        excel_row = DAY_ROW + 1 + row_idx

        row_fill = PatternFill("solid", fgColor="F5F9FF") \
            if row_idx % 2 == 1 else PatternFill("solid", fgColor="FFFFFF")

        def cell(col, val, font=NORMAL_FONT, align=CENTER, fill=row_fill):
            c = ws.cell(excel_row, col, val)
            c.font      = font
            c.alignment = align
            c.border    = BORDER
            c.fill      = fill
            return c

        cell(COL_SR,    row["sr_no"],      font=BOLD_FONT)
        cell(COL_NAME,  row["name"],       font=BOLD_FONT,   align=LEFT)
        cell(COL_DESIG, row["designation"],                  align=LEFT)

        for i, val in enumerate(row["days"]):
            col = COL_D1 + i
            if val == "AB":
                cell(col, "AB",  font=RED_FONT)
            elif val == "WO":
                cell(col, "WO",  font=GREY_FONT)
            else:
                try:
                    num = float(val)
                    is_ot = num > 10.0
                    display = _fmt_decimal(val)
                    cell(col, display, font=RED_FONT if is_ot else NORMAL_FONT)
                except (ValueError, TypeError):
                    cell(col, val)

        c = ws.cell(excel_row, COL_MANDAYS, row["mandays"])
        c.font      = Font(name="Arial", bold=True, color="1565C0", size=9)
        c.fill      = PatternFill("solid", fgColor="E3F2FD")
        c.alignment = CENTER
        c.border    = BORDER

        ot_val = _fmt_decimal(row["ot"]) if row["ot"] else "0"
        c = ws.cell(excel_row, COL_OT, ot_val)
        c.font      = Font(name="Arial", bold=True, color="F57F17", size=9)
        c.fill      = PatternFill("solid", fgColor="FFF8E1")
        c.alignment = CENTER
        c.border    = BORDER

        cell(COL_REMARKS, "")

    ws.column_dimensions[get_column_letter(COL_SR)].width    = 5
    ws.column_dimensions[get_column_letter(COL_NAME)].width  = 22
    ws.column_dimensions[get_column_letter(COL_DESIG)].width = 22
    for i in range(n_days):
        ws.column_dimensions[get_column_letter(COL_D1 + i)].width = 5.5
    ws.column_dimensions[get_column_letter(COL_MANDAYS)].width  = 8
    ws.column_dimensions[get_column_letter(COL_OT)].width       = 7
    ws.column_dimensions[get_column_letter(COL_REMARKS)].width  = 12

    ws.row_dimensions[1].height = 16
    ws.row_dimensions[HDR_ROW].height = 28
    ws.row_dimensions[DAY_ROW].height = 14
    for r in range(DAY_ROW + 1, DAY_ROW + 1 + len(rows)):
        ws.row_dimensions[r].height = 16

    ws.freeze_panes = ws.cell(DAY_ROW + 1, COL_D1)

    return wb


def _fmt_decimal(val):
    """Format a float attendance value for Excel display, e.g. 9.5 → '09.50'"""
    try:
        num = float(val)
        if num <= 0:
            return "0"
        h = int(num)
        m = round((num - h) * 100)
        return f"{h:02d}.{m:02d}"
    except (TypeError, ValueError):
        return str(val) if val else "0"
